from typing import Dict, List, Tuple, Optional, Union
import ast
import openpyxl, os
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment



save_dir = os.getenv('appdata')
file_name: Optional[str] = None
workbook: Optional[Workbook] = None

class CustomComment(Comment):
    def __init__(self, speaker_name: str, src_model: str):
        data = {}
        if self._validateInput(speaker_name): data['speaker_name'] = speaker_name
        if self._validateInput(src_model): data['src_model'] = src_model
        super().__init__(str(data), author= "")

    def _validateInput(self, text: str) -> bool: 
        if isinstance(text, str) and len(text): return True
        return False

    @staticmethod
    def get_speaker_name(text: str): 
        try:
            data = ast.literal_eval(text)
            return data['speaker_name']
        except: return ""


class TextDTO:
    history: Optional[List[str]]
    translatedText: Union[None, str, List[str]]
    def __init__(self, DTO: Dict[str, str]):
        self.window_title = DTO["window_title"]
        self.originalText = DTO["originalText"]
        self.src_model = DTO['src_model'] if 'src_model' in DTO else None
        self.translatedText = DTO["translatedText"] if "translatedText" in DTO else None
        self._speakerName = DTO['speakerName'] if 'speakerName' in DTO else None

    @property
    def speakerName(self):
        return self._speakerName if self._speakerName is not None else ""


def loadFile(name: str): 
    global workbook, file_name
    file_name = name
    try: workbook = openpyxl.load_workbook(filePath())
    except: 
        workbook = Workbook()
        workbook.active.title = 'Translation'

def worksheet() -> Worksheet: 
    if workbook is not None: return workbook['Translation']

def filePath() -> str: return os.path.join(save_dir, 'ai-translate', file_name+'.xlsx')

def queryEntry(textDTO: TextDTO):
    if isinstance(textDTO, TextDTO) and textDTO.window_title is not None:
        if textDTO.window_title != file_name or worksheet() is None: loadFile(textDTO.window_title)
        if worksheet() is not None:
            column: Tuple[Cell] = worksheet()['A']
            for cell in column:
                if cell.value is not None and textDTO.originalText == cell.value: return cell.row


def get_history(lastRowNumber: int):
    if isinstance(lastRowNumber, int) and lastRowNumber is not None:
        history = []
        rowNumber = lastRowNumber
        while len(history) < 10:
            rowNumber = rowNumber-1
            if rowNumber > 0:
                cell = worksheet().cell(row= rowNumber, column=2)
                if cell.value is not None and cell.value != "": 
                    speaker_name = CustomComment.get_speaker_name(cell.comment.text)
                    formated_name = f"[{speaker_name}]: " if len(speaker_name) else ""
                    history.append(formated_name + cell.value)
            else: break

        history.reverse()
        return history



class XLSX:
    def QueryTranslation(self, textDTO: Dict[str, str]) -> TextDTO: 
        textDTO: TextDTO = TextDTO(textDTO)
        entry = queryEntry(textDTO)
        if entry is not None: 
            textDTO.history = get_history(entry)
            textDTO.translatedText = [cell.value for cell in worksheet()[entry][1:]]
            textDTO.translatedText.reverse()
        elif worksheet() is not None: textDTO.history = get_history(worksheet().max_row+1)

        return textDTO.__dict__


    def SaveText(self, textDTO: Dict[str, str]):
        textDTO: TextDTO = TextDTO(textDTO)
        entry = queryEntry(textDTO)
        if entry is not None and worksheet() is not None: 
            row = worksheet()[entry]
            cell = worksheet().cell(row= entry, column= len(row))
            cell.value = textDTO.translatedText
            cell.comment = CustomComment(textDTO.speakerName, textDTO.src_model)

        elif worksheet() is not None: 
            lastRow = worksheet().max_row+1 # same variable for both calls to avoid the "stairs" bug
            cellA = worksheet().cell(row= lastRow, column=1)
            cellB = worksheet().cell(row= lastRow, column=2)
            cellA.value = textDTO.originalText
            cellB.value = textDTO.translatedText
            cellB.comment = CustomComment(textDTO.speakerName, textDTO.src_model)

        else: return

        workbook.save(filePath())
        response = self.QueryTranslation(textDTO.__dict__)
        if response is None:
            return { "error": "failed to save text" }


